from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Literal, Optional

import csv
import re
import json
from collections import OrderedDict
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from paths import DATOS_NEGOCIO_PATH

ClaseDoc = Literal["1", "2", "4"]
TipoDocCF = Literal["01", "02", "10", "11"]


@dataclass
class VentaCF:
    fecha: str
    clase: ClaseDoc
    tipo: TipoDocCF
    numero_resolucion: Optional[str] = None
    serie: Optional[str] = None
    ctrl_interno_del: Optional[str] = None
    ctrl_interno_al: Optional[str] = None
    numero_doc_del: Optional[str] = None
    numero_doc_al: Optional[str] = None
    nro_maquina: Optional[str] = None
    ventas_exentas: str = "0.00"
    internas_exentas_ns: str = "0.00"
    ventas_no_sujetas: str = "0.00"
    ventas_gravadas_locales: str = "0.00"
    exp_ca: str = "0.00"
    exp_fuera_ca: str = "0.00"
    exp_servicios: str = "0.00"
    zonas_francas_dpa: str = "0.00"
    terceros_no_domic: str = "0.00"
    total_ventas: str = "0.00"
    tipo_operacion: str = "0"
    tipo_ingreso: str = "0"
    codigo_generacion: Optional[str] = None
    numero_control: Optional[str] = None
    estado: Optional[str] = None
    estado_manual: Optional[str] = None
    estado_fuente: Optional[str] = None
    estado_documento: Optional[str] = None
    estado_envio: Optional[str] = None
    json_path: Optional[str] = None


HEADERS_CF = [
    "A_FechaEmision",
    "B_Clase",
    "C_Tipo",
    "D_NumResolucion",
    "E_Serie",
    "F_CtrlInterno_Del",
    "G_CtrlInterno_Al",
    "H_NumDoc_Del",
    "I_NumDoc_Al",
    "J_NroMaquina",
    "K_VentasExentas",
    "L_IntExentasNoProp",
    "M_VentasNoSujetas",
    "N_VentasGravadasLocales",
    "O_Exp_CA",
    "P_Exp_Fuera_CA",
    "Q_Exp_Servicios",
    "R_ZonasFrancas_DPA",
    "S_TercerosNoDomic",
    "T_TotalVentas",
    "U_TipoOperacion",
    "V_TipoIngreso",
    "W_NumAnexo",
]

PERIODO_RE = re.compile(r"^\d{6}$")
DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
MONTO_FIELDS = [
    "ventas_exentas",
    "internas_exentas_ns",
    "ventas_no_sujetas",
    "ventas_gravadas_locales",
    "exp_ca",
    "exp_fuera_ca",
    "exp_servicios",
    "zonas_francas_dpa",
    "terceros_no_domic",
    "total_ventas",
]
CERO_DECIMAL = Decimal("0.00")
RENTA_DEFAULT_OPERACION = "1"
RENTA_DEFAULT_INGRESO = "3"
RENTA_DEFAULT_STRICT = True
VALIDATION_STRICT_DEFAULT = True
LENGTH_LIMITS_CF = {
    "numero_resolucion": 20,
    "serie": 20,
    "ctrl_interno_del": 20,
    "ctrl_interno_al": 20,
    "numero_doc_del": 40,
    "numero_doc_al": 40,
    "nro_maquina": 20,
    "codigo_generacion": 40,
    "numero_control": 40,
    "tipo_operacion": 2,
    "tipo_ingreso": 2,
}
DEFAULT_CSV_DELIMITER = ";"

__all__ = [
    "VentaCF",
    "HEADERS_CF",
    "generar_anexo_consumidor_final_files",
    "on_click_generar_consumidor_final",
]


def _validate_periodo(periodo_yyyymm: str) -> str:
    texto = str(periodo_yyyymm).strip()
    if not PERIODO_RE.fullmatch(texto):
        raise ValueError("El período debe tener el formato YYYYMM (6 dígitos).")
    anio = int(texto[:4])
    mes = int(texto[4:])
    if mes < 1 or mes > 12:
        raise ValueError("El período debe contener un mes válido (01-12).")
    if anio < 1900:
        raise ValueError("El año del período es inválido.")
    return texto


def _validate_output_dir(output_dir: str | Path) -> Path:
    text = str(output_dir).strip()
    if not text:
        raise ValueError("Debe indicar una carpeta de salida válida.")
    path = Path(text).expanduser().resolve()
    path.mkdir(parents=True, exist_ok=True)
    return path


def _parse_fecha(fecha: str, periodo: str, idx: int) -> tuple[date, str]:
    valor = (fecha or "").strip()
    if not DATE_RE.fullmatch(valor):
        raise ValueError(
            f"La fecha debe tener formato DD/MM/AAAA (registro {idx + 1})."
        )
    try:
        parsed = datetime.strptime(valor, "%d/%m/%Y").date()
    except ValueError as exc:
        raise ValueError(
            f"Fecha inválida en registro {idx + 1}: {valor!r}."
        ) from exc
    periodo_anio = int(periodo[:4])
    periodo_mes = int(periodo[4:])
    if parsed.year != periodo_anio or parsed.month != periodo_mes:
        raise ValueError(
            "La fecha debe corresponder al período declarado"
            f" (registro {idx + 1})."
        )
    return parsed, parsed.strftime("%d/%m/%Y")


def _normalize_decimal(
    valor: object,
    campo: str,
    idx: int,
    permite_negativo: bool,
) -> Decimal:
    texto = str(valor).strip() if valor is not None else "0"
    if texto == "":
        texto = "0"
    if texto.count(",") > 0:
        # Permitir coma decimal reemplazándola por punto cuando no hay separadores de miles.
        if texto.count(",") == 1 and "." not in texto:
            texto = texto.replace(",", ".")
        else:
            raise ValueError(
                f"Formato numérico inválido en {campo} (registro {idx + 1})."
            )
    if " " in texto:
        raise ValueError(
            f"Formato numérico inválido en {campo} (registro {idx + 1})."
        )
    try:
        monto = Decimal(texto)
    except InvalidOperation as exc:
        raise ValueError(
            f"Valor numérico inválido en {campo} (registro {idx + 1})."
        ) from exc
    if monto < 0 and not permite_negativo:
        raise ValueError(
            f"No se permiten montos negativos en {campo}"
            f" (registro {idx + 1})."
        )
    monto = monto.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return monto


def _permite_negativos(registro: VentaCF) -> bool:
    flag = getattr(registro, "anulado", None)
    if flag is not None:
        return bool(flag)
    flag = getattr(registro, "permite_negativos", None)
    if flag is not None:
        return bool(flag)
    estado = getattr(registro, "estado", None)
    if isinstance(estado, str) and estado.strip().upper() in {
        "A",
        "ANULADO",
        "INVALIDADO",
        "I",
    }:
        return True
    return False


def _validar_total_montos(
    montos: Dict[str, Decimal],
    *,
    contexto: str,
) -> None:
    # Validación desactivada según requisitos actuales del anexo.
    return


def _load_renta_config() -> dict[str, object]:
    cfg = {
        "tipo_operacion_default": RENTA_DEFAULT_OPERACION,
        "tipo_ingreso_default": RENTA_DEFAULT_INGRESO,
        "strict": RENTA_DEFAULT_STRICT,
        "strict_validations": VALIDATION_STRICT_DEFAULT,
        "csv_delimiter": DEFAULT_CSV_DELIMITER,
    }
    try:
        raw = json.loads(Path(DATOS_NEGOCIO_PATH).read_text(encoding="utf-8"))
    except Exception:
        return cfg
    if not isinstance(raw, dict):
        return cfg
    anexos = raw.get("anexos_cf") or raw.get("anexos")
    if not isinstance(anexos, dict):
        return cfg
    op = anexos.get("renta_tipo_operacion_default")
    ing = anexos.get("renta_tipo_ingreso_default")
    strict = anexos.get("strict_renta_fields")
    delim = anexos.get("csv_delimiter")
    if isinstance(op, str) and op.strip():
        cfg["tipo_operacion_default"] = op.strip()
    if isinstance(ing, str) and ing.strip():
        cfg["tipo_ingreso_default"] = ing.strip()
    if isinstance(strict, bool):
        cfg["strict"] = strict
    if isinstance(delim, str) and delim.strip():
        cfg["csv_delimiter"] = delim.strip()
    return cfg


def _resolve_renta_fields(
    registro: VentaCF,
    periodo: str,
    montos: Dict[str, Decimal] | None,
    renta_cfg: dict[str, object] | None = None,
) -> tuple[str, str]:
    cfg = renta_cfg or _load_renta_config()
    if periodo < "202501":
        return "0", "0"

    tipo_operacion = str(registro.tipo_operacion or "").strip()
    tipo_ingreso = str(registro.tipo_ingreso or "").strip()

    if not tipo_operacion:
        if montos and montos.get("ventas_gravadas_locales", CERO_DECIMAL) > CERO_DECIMAL:
            tipo_operacion = "1"
        elif montos:
            tipo_operacion = "2"
        else:
            tipo_operacion = str(cfg.get("tipo_operacion_default") or "").strip()

    if not tipo_ingreso:
        tipo_ingreso = str(cfg.get("tipo_ingreso_default") or "").strip()

    def _validate(valor: str, campo: str) -> str:
        texto = valor.strip()
        if not texto or texto == "0":
            raise ValueError(f"{campo} requerido para periodo >= 2025-01.")
        if not re.fullmatch(r"\d{1,2}", texto):
            raise ValueError(f"{campo} inválido (solo 1-2 dígitos).")
        return texto

    strict = bool(cfg.get("strict", RENTA_DEFAULT_STRICT))
    try:
        tipo_operacion = _validate(tipo_operacion, "Tipo de operación")
    except ValueError:
        if strict:
            raise
        tipo_operacion = str(cfg.get("tipo_operacion_default") or "1").strip() or "1"
        tipo_operacion = _validate(tipo_operacion, "Tipo de operación")

    try:
        tipo_ingreso = _validate(tipo_ingreso, "Tipo de ingreso")
    except ValueError:
        if strict:
            raise
        tipo_ingreso = str(cfg.get("tipo_ingreso_default") or "3").strip() or "3"
        tipo_ingreso = _validate(tipo_ingreso, "Tipo de ingreso")

    return tipo_operacion, tipo_ingreso


def _normalize_renta_fields(registro: VentaCF, periodo: str) -> tuple[str, str]:
    # Compatibilidad: delega a la nueva resolución con montos opcionales.
    return _resolve_renta_fields(registro, periodo, None)


def _validate_lengths_cf(registro: VentaCF, idx: int, strict: bool = True, renta_cfg: dict[str, object] | None = None) -> None:
    limits = LENGTH_LIMITS_CF
    cfg = renta_cfg or _load_renta_config()
    strict_valid = bool(cfg.get("strict_validations", VALIDATION_STRICT_DEFAULT))
    if strict_valid is False:
        strict = False

    def _check(value: str | None, key: str) -> None:
        if value is None:
            return
        texto = str(value).strip()
        if not texto:
            return
        limit = limits.get(key)
        if limit and len(texto) > limit:
            msg = f"{key} excede longitud {limit} (registro {idx + 1})"
            if strict:
                raise ValueError(msg)
            else:
                import logging
                logging.getLogger(__name__).warning(msg)

    _check(registro.numero_resolucion, "numero_resolucion")
    _check(registro.serie, "serie")
    _check(registro.ctrl_interno_del, "ctrl_interno_del")
    _check(registro.ctrl_interno_al, "ctrl_interno_al")
    _check(registro.numero_doc_del, "numero_doc_del")
    _check(registro.numero_doc_al, "numero_doc_al")
    _check(registro.nro_maquina, "nro_maquina")
    _check(registro.codigo_generacion, "codigo_generacion")
    _check(registro.numero_control, "numero_control")
    _check(registro.tipo_operacion, "tipo_operacion")
    _check(registro.tipo_ingreso, "tipo_ingreso")


def _decimal_to_text(monto: Decimal) -> str:
    monto = monto.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{monto:.2f}"


def _autoajustar_columnas(ws) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
        )
        max_length = max(max_length, len(HEADERS_CF[idx - 1]))
        column_letter = get_column_letter(idx)
        ws.column_dimensions[column_letter].width = max_length + 2
        for cell in column_cells:
            cell.number_format = "@"


def _ensure_equals(value_a: str, value_b: str, campo_a: str, campo_b: str, idx: int) -> tuple[str, str]:
    a = (value_a or "").strip()
    b = (value_b or "").strip()
    if a and b and a != b:
        raise ValueError(
            f"{campo_a} y {campo_b} deben coincidir (registro {idx + 1})."
        )
    if not a and not b:
        raise ValueError(
            f"Debe indicar {campo_a} o {campo_b} (registro {idx + 1})."
        )
    if not a:
        a = b
    if not b:
        b = a
    return a, b


def _extract_dte_order(registro: VentaCF, idx: int, fecha: date) -> float:
    candidatos = [
        "orden_emision",
        "orden",
        "sequence",
        "secuencia",
        "timestamp",
        "fecha_hora",
        "fecha_hora_emision",
        "fecha_hora_generacion",
        "hora",
        "hora_emision",
        "created_at",
    ]

    for nombre in candidatos:
        if not hasattr(registro, nombre):
            continue
        valor = getattr(registro, nombre)
        if valor is None:
            continue
        if isinstance(valor, (int, float, Decimal)):
            return float(valor)
        if isinstance(valor, datetime):
            return valor.timestamp()
        if isinstance(valor, date):
            dt = datetime.combine(valor, time.min)
            return dt.timestamp()
        if isinstance(valor, str):
            texto = valor.strip()
            if not texto:
                continue
            try:
                dt = datetime.fromisoformat(texto)
                return dt.timestamp()
            except ValueError:
                pass
            for fmt in (
                "%d/%m/%Y %H:%M:%S",
                "%d/%m/%Y %H:%M",
                "%H:%M:%S",
                "%H:%M",
            ):
                try:
                    dt = datetime.strptime(texto, fmt)
                except ValueError:
                    continue
                if fmt.startswith("%H"):
                    dt = datetime.combine(fecha, dt.time())
                return dt.timestamp()
            if texto.isdigit():
                return float(int(texto))
    return float(idx)


def _validate_output_row(fila: list[str], idx: int, periodo: str) -> None:
    if len(fila) != 23:
        raise ValueError(f"Fila {idx} incompleta: se esperaban 23 columnas.")
    clase = fila[1]
    is_dte = clase == "4"
    if is_dte:
        if fila[3] != "N/A" or fila[4] != "N/A" or fila[5] != "N/A" or fila[6] != "N/A":
            raise ValueError(f"Fila {idx}: para DTE la resolución/serie/control deben ser 'N/A'.")
        if not fila[7] or not fila[8]:
            raise ValueError(f"Fila {idx}: debe indicar código de generación DEL y AL.")
        if fila[9] not in ("", None):
            raise ValueError(f"Fila {idx}: la máquina registradora debe quedar en blanco para DTE.")
        if periodo >= "202501":
            if fila[20] in ("0", "", None) or fila[21] in ("0", "", None):
                raise ValueError(f"Fila {idx}: Tipo de operación/ingreso requeridos para 2025-01 en adelante.")
    # Longitud máxima defensiva (evita rechazos por celdas extensas).
    max_len = 100
    for col_idx, val in enumerate(fila, start=1):
        texto = "" if val is None else str(val)
        if len(texto) > max_len:
            raise ValueError(f"Fila {idx} columna {col_idx} excede {max_len} caracteres.")


def generar_anexo_consumidor_final_files(
    registros: Iterable[VentaCF],
    output_dir: str,
    periodo_yyyymm: str,
) -> Dict[str, Path]:
    periodo = _validate_periodo(periodo_yyyymm)
    output_path = _validate_output_dir(output_dir)
    renta_cfg = _load_renta_config()
    csv_delimiter = renta_cfg.get("csv_delimiter") or DEFAULT_CSV_DELIMITER

    registros_list = list(registros)
    if not registros_list:
        raise ValueError("No hay ventas para generar el Anexo II de consumidor final.")

    filas: list[list[str]] = []
    dte_groups: "OrderedDict[tuple[str, str], dict]" = OrderedDict()

    for idx, registro in enumerate(registros_list):
        if not isinstance(registro, VentaCF):
            raise ValueError(
                f"Cada registro debe ser VentaCF (registro {idx + 1})."
            )
        clase = (registro.clase or "").strip()
        tipo = (registro.tipo or "").strip()
        if clase not in {"1", "2", "4"}:
            raise ValueError(
                f"Clase de documento inválida (registro {idx + 1})."
            )
        if tipo not in {"01", "02", "10", "11"}:
            raise ValueError(
                f"Tipo de documento inválido (registro {idx + 1})."
            )

        parsed_date, fecha_text = _parse_fecha(registro.fecha, periodo, idx)
        permite_neg = _permite_negativos(registro)
        montos: Dict[str, Decimal] = {}
        for field_name in MONTO_FIELDS:
            valor = getattr(registro, field_name)
            montos[field_name] = _normalize_decimal(
                valor,
                field_name,
                idx,
                permite_neg,
            )
        _validar_total_montos(
            montos,
            contexto=f"registro {idx + 1}",
        )
        tipo_operacion, tipo_ingreso = _resolve_renta_fields(
            registro, periodo, montos, renta_cfg
        )

        numero_resolucion = (registro.numero_resolucion or "").strip()
        serie = (registro.serie or "").strip()
        ctrl_del = (registro.ctrl_interno_del or "").strip()
        ctrl_al = (registro.ctrl_interno_al or "").strip()
        doc_del = (registro.numero_doc_del or "").strip()
        doc_al = (registro.numero_doc_al or "").strip()
        nro_maquina = (registro.nro_maquina or "").strip()

        if clase == "4":
            if not doc_del and not doc_al:
                raise ValueError(
                    f"Debe indicar el código de generación del DTE (registro {idx + 1})."
                )
            codigo_inicio = doc_del or doc_al
            codigo_fin = doc_al or doc_del
            key = (fecha_text, tipo)
            if key not in dte_groups:
                dte_groups[key] = {
                    "fecha": parsed_date,
                    "fecha_text": fecha_text,
                    "tipo": tipo,
                    "montos": {name: CERO_DECIMAL for name in MONTO_FIELDS},
                    "codigos": [],
                    "tipo_operacion": tipo_operacion,
                    "tipo_ingreso": tipo_ingreso,
                }
            group = dte_groups[key]
            group_montos = group["montos"]
            for name, amount in montos.items():
                group_montos[name] = group_montos[name] + amount
            orden = _extract_dte_order(registro, idx, parsed_date)
            group["codigos"].append(
                {
                    "inicio": codigo_inicio,
                    "fin": codigo_fin,
                    "orden": orden,
                }
            )
            if group["tipo_operacion"] != tipo_operacion:
                raise ValueError(
                    "Tipo de operación inconsistente dentro del mismo día/tipo."
                )
            if group["tipo_ingreso"] != tipo_ingreso:
                raise ValueError(
                    "Tipo de ingreso inconsistente dentro del mismo día/tipo."
                )
            continue

        ctrl_del, doc_del = _ensure_equals(ctrl_del, doc_del, "Control interno (DEL)", "Número de documento (DEL)", idx)
        ctrl_al, doc_al = _ensure_equals(ctrl_al, doc_al, "Control interno (AL)", "Número de documento (AL)", idx)

        if clase in {"1", "2"} and tipo == "10" and not nro_maquina:
            raise ValueError(
                f"Debe indicar el número de máquina registradora (registro {idx + 1})."
            )

        _validate_lengths_cf(
            registro,
            idx,
            strict=True,
            renta_cfg=renta_cfg,
        )
        fila = [
            fecha_text,
            clase,
            tipo,
            numero_resolucion,
            serie,
            ctrl_del,
            ctrl_al,
            doc_del,
            doc_al,
            nro_maquina,
        ]
        fila.extend(_decimal_to_text(montos[name]) for name in MONTO_FIELDS)
        fila.extend([tipo_operacion, tipo_ingreso, "2"])
        filas.append(fila)

    for key in dte_groups:
        group = dte_groups[key]
        codigos = group["codigos"]
        codigos_ordenados = sorted(
            codigos,
            key=lambda item: (item["orden"], item["inicio"]),
        )
        codigo_inicio = codigos_ordenados[0]["inicio"]
        codigo_fin = codigos_ordenados[-1]["fin"]
        _validar_total_montos(
            group["montos"],
            contexto=f"DTE del {group['fecha_text']} tipo {group['tipo']}",
        )
        dummy_reg = VentaCF(
            fecha=group["fecha_text"],
            clase="4",
            tipo=group["tipo"],
            numero_doc_del=codigo_inicio,
            numero_doc_al=codigo_fin,
            tipo_operacion=group["tipo_operacion"],
            tipo_ingreso=group["tipo_ingreso"],
            codigo_generacion=codigo_inicio,
            numero_control=None,
        )
        _validate_lengths_cf(
            dummy_reg,
            len(filas) + len(dte_groups),
            strict=True,
            renta_cfg=renta_cfg,
        )
        fila = [
            group["fecha_text"],
            "4",
            group["tipo"],
            "N/A",
            "N/A",
            "N/A",
            "N/A",
            codigo_inicio,
            codigo_fin,
            "",
        ]
        fila.extend(
            _decimal_to_text(group["montos"][name]) for name in MONTO_FIELDS
        )
        fila.extend([group["tipo_operacion"], group["tipo_ingreso"], "2"])
        filas.append(fila)

    # Ordenar por fecha y tipo para mantener consistencia en la salida.
    filas.sort(key=lambda row: (datetime.strptime(row[0], "%d/%m/%Y"), row[2], row[1]))

    csv_name = f"Anexo_II_Consumidor_Final_{periodo}.csv"
    xlsx_name = f"Anexo_II_Consumidor_Final_{periodo}.xlsx"
    csv_path = output_path / csv_name
    xlsx_path = output_path / xlsx_name

    with open(csv_path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(
            fh,
            delimiter=csv_delimiter,
            lineterminator="\n",
            quoting=csv.QUOTE_MINIMAL,
        )
        for idx, fila in enumerate(filas, start=1):
            _validate_output_row(fila, idx, periodo)
            writer.writerow(fila)

    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo II"
    ws.append(HEADERS_CF)
    for fila in filas:
        ws.append(fila)
    _autoajustar_columnas(ws)
    wb.save(xlsx_path)

    return {"csv": csv_path, "xlsx": xlsx_path}


def on_click_generar_consumidor_final(
    output_dir: str,
    periodo_yyyymm: str,
    registros: Iterable[VentaCF],
) -> Dict[str, object]:
    registros_list = list(registros)

    try:
        paths = generar_anexo_consumidor_final_files(
            registros_list, output_dir, periodo_yyyymm
        )
    except ValueError as exc:
        return {"success": False, "message": str(exc), "paths": {}}
    except Exception as exc:
        return {
            "success": False,
            "message": f"Error inesperado: {exc}",
            "paths": {},
        }

    csv_path = paths["csv"]
    xlsx_path = paths["xlsx"]
    cantidad = len(registros_list)
    message_lines = [
        f"{cantidad} DTE exportados al Anexo II.",
        f"CSV: {csv_path}",
        f"XLSX: {xlsx_path}",
    ]
    return {
        "success": True,
        "message": "\n".join(message_lines),
        "paths": paths,
        "count": cantidad,
    }
