from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from declaracion.anexo_consumidor_final import (
    HEADERS_CF,
    VentaCF,
    generar_anexo_consumidor_final_files,
)


def _read_csv(path: Path) -> list[list[str]]:
    content = path.read_text(encoding="utf-8")
    lines = [line for line in content.splitlines() if line]
    return [line.split(";") for line in lines]


def test_generar_dte_agrupado_por_dia(tmp_path):
    registros = [
        VentaCF(
            fecha="05/01/2025",
            clase="4",
            tipo="01",
            numero_doc_del="0000002-BBB",
            numero_doc_al="0000002-BBB",
            ventas_gravadas_locales="12.25",
            total_ventas="12.25",
            tipo_operacion="1",
            tipo_ingreso="2",
        ),
        VentaCF(
            fecha="05/01/2025",
            clase="4",
            tipo="01",
            numero_doc_del="0000001-AAA",
            numero_doc_al="0000001-AAA",
            ventas_gravadas_locales="10.50",
            total_ventas="10.50",
            tipo_operacion="1",
            tipo_ingreso="2",
        ),
        VentaCF(
            fecha="05/01/2025",
            clase="4",
            tipo="01",
            numero_doc_del="0000003-CCC",
            numero_doc_al="0000003-CCC",
            ventas_gravadas_locales="5.00",
            total_ventas="5.00",
            tipo_operacion="1",
            tipo_ingreso="2",
        ),
    ]

    registros[0].orden_emision = 2
    registros[1].orden_emision = 1
    registros[2].orden_emision = 3

    paths = generar_anexo_consumidor_final_files(registros, tmp_path, "202501")

    rows = _read_csv(paths["csv"])
    assert len(rows) == 1
    row = rows[0]
    assert len(row) == 23
    assert row[0] == "05/01/2025"
    assert row[1] == "4"
    assert row[3] == "N/A" and row[4] == "N/A"
    assert row[5] == "N/A" and row[6] == "N/A"
    assert row[7] == "0000001-AAA"
    assert row[8] == "0000003-CCC"
    assert row[9] == ""
    assert Decimal(row[13]) == Decimal("27.75")
    assert Decimal(row[19]) == Decimal("27.75")
    assert row[20] == "1"
    assert row[21] == "2"
    assert row[22] == "2"

    csv_content = paths["csv"].read_text(encoding="utf-8")
    assert csv_content.endswith("\n")
    assert ";".join(row) + "\n" == csv_content

    wb = load_workbook(paths["xlsx"])
    ws = wb.active
    assert [cell.value for cell in ws[1]] == HEADERS_CF
    assert ws[2][7].value == "0000001-AAA"


def test_generar_impreso_ctrl_y_rangos(tmp_path):
    registro = VentaCF(
        fecha="15/02/2024",
        clase="1",
        tipo="10",
        numero_resolucion="123",
        serie="A1",
        ctrl_interno_del="0001",
        ctrl_interno_al="0001",
        numero_doc_del="0001",
        numero_doc_al="0001",
        nro_maquina="MR-01",
        ventas_gravadas_locales="8.00",
        total_ventas="8.00",
    )

    paths = generar_anexo_consumidor_final_files([registro], tmp_path, "202402")

    row = _read_csv(paths["csv"])[0]
    assert row[5] == row[7] == "0001"
    assert row[6] == row[8] == "0001"
    assert row[9] == "MR-01"
    assert row[20] == "0"
    assert row[21] == "0"
    assert row[22] == "2"


def test_formato_validaciones_y_renta(tmp_path):
    registro_valido = VentaCF(
        fecha="10/01/2024",
        clase="2",
        tipo="01",
        ctrl_interno_del="100",
        ctrl_interno_al="100",
        numero_doc_del="100",
        numero_doc_al="100",
        ventas_gravadas_locales="1.235",
        total_ventas="1.235",
    )

    paths = generar_anexo_consumidor_final_files([registro_valido], tmp_path, "202401")
    row = _read_csv(paths["csv"])[0]
    assert row[13] == "1.24"
    assert row[19] == "1.24"
    assert row[20] == "0"
    assert row[21] == "0"
    assert row[22] == "2"
    assert row.count(";") == 0  # split removed separators
    csv_line = paths["csv"].read_text(encoding="utf-8").splitlines()[0]
    assert csv_line.count(";") == 22
    assert not csv_line.startswith("A_FechaEmision")

    wb = load_workbook(paths["xlsx"])
    ws = wb.active
    assert [cell.value for cell in ws[1]] == HEADERS_CF

    registro_fuera_periodo = VentaCF(
        fecha="05/02/2024",
        clase="1",
        tipo="01",
        ctrl_interno_del="1",
        ctrl_interno_al="1",
        numero_doc_del="1",
        numero_doc_al="1",
    )

    with pytest.raises(ValueError):
        generar_anexo_consumidor_final_files([registro_fuera_periodo], tmp_path, "202401")


def test_csv_tiene_23_columnas_por_fila(tmp_path):
    registros = [
        VentaCF(
            fecha="01/03/2025",
            clase="4",
            tipo="01",
            numero_doc_del="A-1",
            numero_doc_al="A-1",
            ventas_gravadas_locales="3",
            total_ventas="3",
            tipo_operacion="1",
            tipo_ingreso="2",
        ),
        VentaCF(
            fecha="02/03/2025",
            clase="1",
            tipo="02",
            ctrl_interno_del="10",
            ctrl_interno_al="10",
            numero_doc_del="10",
            numero_doc_al="10",
            ventas_exentas="1.50",
            ventas_gravadas_locales="2.50",
            total_ventas="4.00",
        ),
    ]

    paths = generar_anexo_consumidor_final_files(registros, tmp_path, "202503")

    for line in paths["csv"].read_text(encoding="utf-8").splitlines():
        assert line.count(";") == 22


def test_permite_negativos_segun_flag(tmp_path):
    registro_negativo = VentaCF(
        fecha="05/04/2024",
        clase="2",
        tipo="01",
        ctrl_interno_del="1",
        ctrl_interno_al="1",
        numero_doc_del="1",
        numero_doc_al="1",
        ventas_gravadas_locales="-5.00",
        total_ventas="-5.00",
    )
    registro_negativo.permite_negativos = True

    paths = generar_anexo_consumidor_final_files([registro_negativo], tmp_path, "202404")
    row = _read_csv(paths["csv"])[0]
    assert row[13] == "-5.00"
    assert row[19] == "-5.00"

    registro_no_permitido = VentaCF(
        fecha="06/04/2024",
        clase="2",
        tipo="01",
        ctrl_interno_del="2",
        ctrl_interno_al="2",
        numero_doc_del="2",
        numero_doc_al="2",
        ventas_gravadas_locales="-1.00",
        total_ventas="-1.00",
    )

    with pytest.raises(ValueError):
        generar_anexo_consumidor_final_files([registro_no_permitido], tmp_path, "202404")


def test_rangos_tiquete_requiere_maquina(tmp_path):
    registro = VentaCF(
        fecha="07/05/2024",
        clase="1",
        tipo="10",
        ctrl_interno_del="100",
        ctrl_interno_al="100",
        numero_doc_del="100",
        numero_doc_al="100",
        ventas_gravadas_locales="2.00",
        total_ventas="2.00",
    )

    with pytest.raises(ValueError):
        generar_anexo_consumidor_final_files([registro], tmp_path, "202405")


def test_totales_deben_coincidir(tmp_path):
    registro = VentaCF(
        fecha="08/06/2024",
        clase="2",
        tipo="01",
        ctrl_interno_del="200",
        ctrl_interno_al="200",
        numero_doc_del="200",
        numero_doc_al="200",
        ventas_exentas="1.00",
        total_ventas="2.00",
    )

    with pytest.raises(ValueError):
        generar_anexo_consumidor_final_files([registro], tmp_path, "202406")


def test_campos_renta_por_periodo(tmp_path):
    registro_ant = VentaCF(
        fecha="09/07/2024",
        clase="2",
        tipo="01",
        ctrl_interno_del="1",
        ctrl_interno_al="1",
        numero_doc_del="1",
        numero_doc_al="1",
        tipo_operacion="5",
        tipo_ingreso="9",
        ventas_gravadas_locales="1.00",
        total_ventas="1.00",
    )

    row_ant = _read_csv(
        generar_anexo_consumidor_final_files([registro_ant], tmp_path, "202407")["csv"]
    )[0]
    assert row_ant[20] == "0"
    assert row_ant[21] == "0"

    registro_post = VentaCF(
        fecha="10/01/2025",
        clase="2",
        tipo="01",
        ctrl_interno_del="2",
        ctrl_interno_al="2",
        numero_doc_del="2",
        numero_doc_al="2",
        tipo_operacion="5",
        tipo_ingreso="9",
        ventas_gravadas_locales="1.50",
        total_ventas="1.50",
    )

    row_post = _read_csv(
        generar_anexo_consumidor_final_files([registro_post], tmp_path, "202501")["csv"]
    )[0]
    assert row_post[20] == "5"
    assert row_post[21] == "9"
