from pathlib import Path

import pytest
from openpyxl import load_workbook

from declaracion.anexo_contribuyentes import (
    HEADERS,
    VentaContribuyente,
    generar_anexo_contribuyentes_files,
    on_click_generar_contribuyentes,
)


def _leer_csv(path: Path) -> list[list[str]]:
    contenido = path.read_text(encoding="utf-8")
    lineas = [linea for linea in contenido.splitlines() if linea]
    return [linea.split(";") for linea in lineas]


def _ejemplo(**overrides) -> VentaContribuyente:
    data = dict(
        fecha_emision="15/03/2025",
        clase="4",
        tipo="03",
        numero_control="DTE-03-S001-00000000000123",
        codigo_generacion="12345678-1234-4234-8234-1234567890AB",
        sello_recepcion="SELLO1234567890",
        identificacion="0614-290790-102-3",
        nombre_cliente="Empresa ABC, S.A. de C.V.",
        ventas_exentas="0",
        ventas_no_sujetas="0",
        ventas_gravadas_locales="100.00",
        debito_fiscal="13.00",
        ventas_terceros_no_domiciliados="0",
        debito_terceros="0",
        total_ventas="113.00",
        dui=None,
        tipo_operacion="7",
        tipo_ingreso="4",
    )
    data.update(overrides)
    return VentaContribuyente(**data)


def test_generar_anexo_contribuyentes_files(tmp_path: Path) -> None:
    registros = [
        _ejemplo(tipo="03", numero_control="DTE-03-001-00001"),
        _ejemplo(
            tipo="05",
            numero_control="DTE-05-001-00002",
            codigo_generacion="AAAAAAAA-1111-2222-3333-AAAAAAAAAAAA",
            ventas_exentas="5.50",
            ventas_gravadas_locales="0",
            debito_fiscal="0",
            total_ventas="5.50",
        ),
        _ejemplo(
            tipo="06",
            numero_control="DTE-06-001-00003",
            codigo_generacion="BBBBBBBB-1111-2222-3333-BBBBBBBBBBBB",
            ventas_no_sujetas="2.00",
            ventas_gravadas_locales="3.00",
            debito_fiscal="0.39",
            total_ventas="5.39",
            identificacion="0614-290790-102-3",
        ),
    ]

    resultado = generar_anexo_contribuyentes_files(registros, tmp_path, "202503")

    csv_path = resultado["csv"]
    xlsx_path = resultado["xlsx"]

    assert csv_path.name == "Ventas_Contribuyentes_202503.csv"
    assert xlsx_path.name == "Ventas_Contribuyentes_202503.xlsx"

    filas = _leer_csv(csv_path)
    assert len(filas) == 3
    for fila in filas:
        assert len(fila) == 20
        assert fila[1] == "4"
        assert fila[6] == ""
        assert not fila[7].count("-")
        assert fila[9].count(",") == 0
        assert fila[19] == "1"

    assert filas[0][3] == "DTE0300100001"
    assert filas[0][5] == registros[0].codigo_generacion.replace("-", "")
    assert filas[0][17] == "7"
    assert filas[0][18] == "4"

    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert [cell.value for cell in ws[1]] == HEADERS
    assert ws.title == "Anexo I"


def test_registro_pre_noviembre_2022(tmp_path: Path) -> None:
    registro = _ejemplo(
        fecha_emision="10/10/2022",
        numero_control="DTE-03-XYZ-00000000004567",
        codigo_generacion="ABCD1234-1111-2222-3333-ABCDEFABCDEF",
        tipo_operacion="2",
        tipo_ingreso="3",
        total_ventas="113.00",
    )

    filas = _leer_csv(
        generar_anexo_contribuyentes_files([registro], tmp_path, "202210")["csv"]
    )

    assert filas[0][3] == "ABCD1234111122223333ABCDEFABCDEF"
    assert filas[0][5] == "DTE03XYZ00000000004567"
    assert filas[0][17] == "0"
    assert filas[0][18] == "0"


def test_validaciones(tmp_path: Path) -> None:
    invalido = _ejemplo(total_ventas="100.00", debito_fiscal="10.00")
    with pytest.raises(ValueError):
        generar_anexo_contribuyentes_files([invalido], tmp_path, "202503")

    sin_identificacion = _ejemplo(identificacion="", dui="")
    with pytest.raises(ValueError):
        generar_anexo_contribuyentes_files([sin_identificacion], tmp_path, "202503")

    ambos = _ejemplo(identificacion="0614-290790-102-3", dui="012345678")
    with pytest.raises(ValueError):
        generar_anexo_contribuyentes_files([ambos], tmp_path, "202503")

    with pytest.raises(ValueError):
        generar_anexo_contribuyentes_files([], tmp_path, "202503")

    sin_sello = _ejemplo(sello_recepcion="")
    generar_anexo_contribuyentes_files([sin_sello], tmp_path, "202503")


def test_on_click(tmp_path: Path) -> None:
    registros = [_ejemplo()]
    resultado = on_click_generar_contribuyentes(tmp_path, "202503", registros)

    assert resultado["success"] is True
    assert resultado["count"] == 1
    assert "Ventas_Contribuyentes_202503" in resultado["message"]

    fallo = on_click_generar_contribuyentes(tmp_path, "202503", [])
    assert fallo["success"] is False
    assert fallo["count"] == 0
