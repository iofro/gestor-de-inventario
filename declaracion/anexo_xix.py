from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Literal

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:  # pragma: no cover - depende de disponibilidad del catálogo en runtime
    from utils.catalogos import DTE_TIPOS as _DTE_TIPOS
except Exception:  # pragma: no cover - manejado con catálogo por defecto
    _DTE_TIPOS = None

EstadoDetalle = Literal["A", "D", "X"]  # A=anulado, D=invalidado, X=extraviado


@dataclass
class DTEAnulado:
    numero_control: str        # Ej. 'DTE-03-S001P001-000000000000454' (con guiones)
    tipo_documento: str        # '01','02','03','04','05','06','10','11', etc.
    sello_recepcion: str       # 40 chars
    codigo_generacion: str     # 36 chars con guiones
    estado: EstadoDetalle      # 'A'|'D'|'X'


UUID36_RE = re.compile(r"^[0-9A-F]{8}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{12}$", re.IGNORECASE)
SELLO40_RE = re.compile(r"^[0-9A-Z]{40}$")
NUMERO_CONTROL_RE = re.compile(r"^DTE-\d{2}-[A-Z0-9]{8,10}-\d{15}$")
PERIODO_RE = re.compile(r"^\d{6}$")

HEADERS = [
    "A_NumResolucionControl",
    "B_Clase",
    "C_DesdePreimpreso",
    "D_HastaPreimpreso",
    "E_TipoDocumento",
    "F_TipoDetalle",
    "G_Serie",
    "H_DesdeNumero",
    "I_HastaNumero",
    "J_CodigoGeneracion",
]

__all__ = [
    "DTEAnulado",
    "EstadoDetalle",
    "HEADERS",
    "generar_anexo_xix_files",
    "on_click_generar_anulaciones",
]

DEFAULT_TIPOS_VALIDOS = {
    "01",
    "02",
    "03",
    "04",
    "05",
    "06",
    "07",
    "08",
    "09",
    "10",
    "11",
    "12",
    "13",
    "14",
    "15",
}


def _build_catalogo_tipos() -> set[str]:
    if isinstance(_DTE_TIPOS, dict):
        catalogo: set[str] = set()
        for key in _DTE_TIPOS.keys():
            text = str(key).strip().upper()
            catalogo.add(text.zfill(2) if text.isdigit() and len(text) < 2 else text)
        catalogo.update(DEFAULT_TIPOS_VALIDOS)
        return catalogo
    return DEFAULT_TIPOS_VALIDOS


CATALOGO_TIPOS_VALIDOS = _build_catalogo_tipos()


def _validate_periodo(periodo_yyyymm: str) -> str:
    periodo = str(periodo_yyyymm).strip()
    if not PERIODO_RE.fullmatch(periodo):
        raise ValueError("El período debe tener el formato YYYYMM (6 dígitos).")
    return periodo


def _validate_output_dir(output_dir: str | Path) -> Path:
    text = str(output_dir).strip()
    if not text:
        raise ValueError("Debe indicar una carpeta de salida válida.")
    path = Path(text).expanduser().resolve()
    path.mkdir(parents=True, exist_ok=True)
    return path


def _pad2(s: str) -> str:
    s = (s or "").strip().upper()
    return s.zfill(2) if s.isdigit() and len(s) < 2 else s


def _validate_registro(registro: DTEAnulado, idx: int | None = None) -> list[str]:
    idx_info = f" (fila {idx + 1})" if idx is not None else ""
    if not isinstance(registro, DTEAnulado):
        raise ValueError(f"Cada registro debe ser DTEAnulado{idx_info}.")

    numero_control = (registro.numero_control or "").strip().upper()
    if not NUMERO_CONTROL_RE.fullmatch(numero_control):
        raise ValueError(
            "Formato inválido de número de control"
            f"{idx_info}: esperado DTE-XX-XXXXXXXX[-..]-000000000000000."
        )

    tipo_documento = _pad2(registro.tipo_documento)
    if tipo_documento not in CATALOGO_TIPOS_VALIDOS:
        raise ValueError(
            f"Tipo de documento fuera de catálogo{idx_info}: {tipo_documento!r}"
        )

    estado = (registro.estado or "").strip().upper()
    if estado not in {"A", "D", "X"}:
        raise ValueError(f"Estado inválido{idx_info}: {estado!r} (A/D/X)")

    sello_recepcion = (registro.sello_recepcion or "").strip().upper()
    if not SELLO40_RE.fullmatch(sello_recepcion):
        raise ValueError(
            f"Sello de recepción inválido{idx_info}: se esperan 40 A–Z/0–9"
        )

    codigo_generacion = (registro.codigo_generacion or "").strip().upper()
    if not UUID36_RE.fullmatch(codigo_generacion):
        raise ValueError(
            f"Código de generación inválido{idx_info}: UUID de 36 con guiones"
        )

    return [
        numero_control,
        "4",
        "0",
        "0",
        tipo_documento,
        estado,
        sello_recepcion,
        "0",
        "0",
        codigo_generacion,
    ]


def _autoajustar_columnas(ws) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        max_length = max(max_length, len(HEADERS[idx - 1]))
        column_letter = get_column_letter(idx)
        ws.column_dimensions[column_letter].width = max_length + 2
        for cell in column_cells:
            cell.number_format = "@"


def generar_anexo_xix_files(
    registros: Iterable[DTEAnulado], output_dir: str, periodo_yyyymm: str
) -> Dict[str, Path]:
    """
    Genera:
      - XLSX con encabezados (para usuario)
      - CSV Hacienda (sin encabezados, ;, UTF-8, \n)
    Orden de columnas A→J, ver mapeo abajo.
    Valida longitudes y formatos; si algo falla, levantar ValueError con mensaje claro.
    Retorna {'csv': Path, 'xlsx': Path}.
    """

    periodo = _validate_periodo(periodo_yyyymm)
    output_path = _validate_output_dir(output_dir)

    registros_list = list(registros)
    if not registros_list:
        raise ValueError("No hay anulaciones para generar el Anexo XIX.")

    filas = [_validate_registro(registro, i) for i, registro in enumerate(registros_list)]

    csv_name = f"Anexo_XIX_Anulados_{periodo}.csv"
    xlsx_name = f"Anexo_XIX_Anulados_{periodo}.xlsx"

    csv_path = output_path / csv_name
    xlsx_path = output_path / xlsx_name

    with open(csv_path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(
            fh,
            delimiter=";",
            lineterminator="\n",
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writerows(filas)

    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo XIX"
    ws.append(HEADERS)
    for row in filas:
        ws.append(row)
    _autoajustar_columnas(ws)
    wb.save(xlsx_path)

    return {"csv": csv_path, "xlsx": xlsx_path}


def on_click_generar_anulaciones(
    output_dir: str, periodo_yyyymm: str, registros: Iterable[DTEAnulado]
) -> Dict[str, object]:
    """
    Llama a generar_anexo_xix_files, muestra rutas generadas y captura errores.
    """

    try:
        paths = generar_anexo_xix_files(registros, output_dir, periodo_yyyymm)
    except ValueError as exc:  # errores esperados por validación
        return {"success": False, "message": str(exc), "paths": {}}
    except Exception as exc:  # errores inesperados
        return {"success": False, "message": f"Error inesperado: {exc}", "paths": {}}

    csv_path = paths["csv"]
    xlsx_path = paths["xlsx"]
    message_lines = [
        "Archivos generados correctamente:",
        f"CSV: {csv_path}",
        f"XLSX: {xlsx_path}",
    ]
    return {"success": True, "message": "\n".join(message_lines), "paths": paths}
