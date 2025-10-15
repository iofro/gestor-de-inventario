from pathlib import Path

import pytest
from openpyxl import load_workbook

from declaracion.anexo_xix import (
    DTEAnulado,
    HEADERS,
    generar_anexo_xix_files,
    on_click_generar_anulaciones,
)


def _build_example(**overrides) -> DTEAnulado:
    data = dict(
        numero_control="DTE-03-S001P001-000000000000454",
        tipo_documento="03",
        sello_recepcion="20254498066A4AC844F5BFF3A77CF1397CCBB8RG",
        codigo_generacion="5B5FF2DC-3250-4CBB-AB34-F0FC44A29557",
        estado="D",
    )
    data.update(overrides)
    return DTEAnulado(**data)


def test_generar_anexo_xix_files(tmp_path: Path) -> None:
    registros = [_build_example()]
    resultado = generar_anexo_xix_files(registros, str(tmp_path), "202510")

    csv_path = resultado["csv"]
    xlsx_path = resultado["xlsx"]

    assert csv_path.exists()
    assert xlsx_path.exists()
    assert csv_path.name == "Anexo_XIX_Anulados_202510.csv"
    assert xlsx_path.name == "Anexo_XIX_Anulados_202510.xlsx"

    contenido = csv_path.read_bytes()
    assert contenido.count(b";") == 9
    assert b"\r\n" not in contenido
    lineas = contenido.decode("utf-8").strip().split("\n")
    assert len(lineas) == 1
    assert not lineas[0].startswith("A_NumResolucionControl")

    wb = load_workbook(xlsx_path)
    hoja = wb.active
    assert hoja.max_column == 10
    assert [celda.value for celda in hoja[1]] == HEADERS
    assert [celda.value for celda in hoja[2]] == [
        "DTE-03-S001P001-000000000000454",
        "4",
        "0",
        "0",
        "03",
        "D",
        "20254498066A4AC844F5BFF3A77CF1397CCBB8RG",
        "0",
        "0",
        "5B5FF2DC-3250-4CBB-AB34-F0FC44A29557",
    ]


def test_validaciones_anexo_xix(tmp_path: Path) -> None:
    with pytest.raises(ValueError) as sello_err:
        generar_anexo_xix_files([
            _build_example(sello_recepcion="X" * 39)
        ], str(tmp_path), "202510")
    assert "fila 1" in str(sello_err.value)

    with pytest.raises(ValueError) as uuid_err:
        generar_anexo_xix_files([
            _build_example(codigo_generacion="12345678-1234-1234-1234-1234567890")
        ], str(tmp_path), "202510")
    assert "fila 1" in str(uuid_err.value)

    with pytest.raises(ValueError):
        generar_anexo_xix_files([
            _build_example(estado="Z")
        ], str(tmp_path), "202510")

    with pytest.raises(ValueError):
        generar_anexo_xix_files([
            _build_example(tipo_documento="99")
        ], str(tmp_path), "202510")

    with pytest.raises(ValueError):
        generar_anexo_xix_files([
            _build_example()
        ], str(tmp_path), "20251")

    with pytest.raises(ValueError):
        generar_anexo_xix_files([], str(tmp_path), "202510")


def test_on_click_generar_anulaciones(tmp_path: Path) -> None:
    registros = [_build_example()]
    resultado = on_click_generar_anulaciones(str(tmp_path), "202510", registros)
    assert resultado["success"] is True
    assert "Anexo_XIX_Anulados_202510" in resultado["message"]

    fallo = on_click_generar_anulaciones(str(tmp_path), "202510", [])
    assert fallo["success"] is False
    assert "No hay anulaciones" in fallo["message"]


def test_multiple_registros_y_normalizacion(tmp_path: Path) -> None:
    registros = [
        _build_example(),
        _build_example(
            numero_control="DTE-03-S001P001P0-000000000000123",
            tipo_documento="3",
            estado="A",
            codigo_generacion="AAAAAAAA-1111-2222-3333-AAAAAAAAAAAA",
            sello_recepcion="1234567890123456789012345678901234567890",
        ),
        _build_example(
            numero_control="DTE-05-ABCD1234-000000000000999",
            tipo_documento="10",
            estado="X",
            codigo_generacion="BBBBBBBB-1111-2222-3333-BBBBBBBBBBBB",
            sello_recepcion="ABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890ABCD",
        ),
    ]

    resultado = generar_anexo_xix_files(registros, str(tmp_path), "202512")
    contenido = resultado["csv"].read_text(encoding="utf-8").strip().split("\n")
    assert len(contenido) == 3
    segunda = contenido[1].split(";")
    tercera = contenido[2].split(";")
    assert segunda[4] == "03"  # normalizado a dos dígitos
    assert segunda[5] == "A"
    assert tercera[5] == "X"


def test_error_indica_fila(tmp_path: Path) -> None:
    registros = [
        _build_example(),
        _build_example(codigo_generacion="INVALIDO"),
    ]

    with pytest.raises(ValueError) as exc:
        generar_anexo_xix_files(registros, str(tmp_path), "202510")

    assert "fila 2" in str(exc.value)
