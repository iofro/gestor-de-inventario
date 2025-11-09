"""Generación del Anexo I (Ventas a contribuyentes)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, Iterable

import csv
import re
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from declaracion.anexo_consumidor_final import (
    _decimal_to_text,
    _normalize_decimal,
    _normalize_renta_fields,
    _parse_fecha,
    _validate_output_dir,
    _validate_periodo,
)

__all__ = [
    "HEADERS",
    "VentaContribuyente",
    "generar_anexo_contribuyentes_files",
    "on_click_generar_contribuyentes",
]


HEADERS = [
    "A_Fecha",
    "B_Clase",
    "C_Tipo",
    "D_NumResolucion",
    "E_NumSerie",
    "F_NumDocumento",
    "G_NumCtrlInterno",
    "H_NIT_NRC",
    "I_NombreRazonSocial",
    "J_VentasExentas",
    "K_VentasNoSujetas",
    "L_VentasGravadasLocales",
    "M_DebitoFiscal",
    "N_VentasTercerosNoDomic",
    "O_DebitoFiscalTerceros",
    "P_TotalVentas",
    "Q_DUI",
    "R_TipoOperacion",
    "S_TipoIngreso",
    "T_NumAnexo",
]


MONTO_FIELDS = (
    "ventas_exentas",
    "ventas_no_sujetas",
    "ventas_gravadas_locales",
    "debito_fiscal",
    "ventas_terceros_no_domiciliados",
    "debito_terceros",
    "total_ventas",
)


@dataclass
class VentaContribuyente:
    fecha_emision: str
    clase: str
    tipo: str
    numero_control: str | None
    codigo_generacion: str | None
    sello_recepcion: str | None
    identificacion: str | None
    nombre_cliente: str
    ventas_exentas: object = "0"
    ventas_no_sujetas: object = "0"
    ventas_gravadas_locales: object = "0"
    debito_fiscal: object = "0"
    ventas_terceros_no_domiciliados: object = "0"
    debito_terceros: object = "0"
    total_ventas: object = "0"
    dui: str | None = None
    tipo_operacion: str | None = None
    tipo_ingreso: str | None = None
    estado: str | None = None
    estado_manual: str | None = None
    estado_fuente: str | None = None
    estado_documento: str | None = None
    estado_envio: str | None = None
    json_path: str | None = None


_WHITESPACE_RE = re.compile(r"\s+")


def _limpiar_guiones(texto: str) -> str:
    return re.sub(r"[-\s]", "", texto)


def _normalizar_identificacion(valor: str) -> str:
    limpio = _limpiar_guiones(valor)
    return limpio.strip()


def _validar_montos(montos: Dict[str, Decimal], idx: int) -> None:
    # Validación desactivada según requisitos actuales del anexo.
    return


def _autoajustar_columnas(ws) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        max_length = max(max_length, len(HEADERS[idx - 1]))
        column_letter = get_column_letter(idx)
        ws.column_dimensions[column_letter].width = max_length + 2
        for cell in column_cells:
            cell.number_format = "@"


def generar_anexo_contribuyentes_files(
    registros: Iterable[VentaContribuyente],
    output_dir: str | Path,
    periodo_yyyymm: str,
) -> Dict[str, Path]:
    periodo = _validate_periodo(periodo_yyyymm)
    output_path = _validate_output_dir(output_dir)

    registros_list = list(registros)
    if not registros_list:
        raise ValueError("No hay ventas para generar el Anexo I de contribuyentes.")

    filas: list[tuple[tuple, list[str]]] = []

    for idx, registro in enumerate(registros_list):
        if not isinstance(registro, VentaContribuyente):
            raise ValueError(
                f"Cada registro debe ser VentaContribuyente (registro {idx + 1})."
            )

        clase = str(registro.clase or "").strip()
        if clase != "4":
            raise ValueError(
                f"Clase de documento inválida (registro {idx + 1})."
            )

        tipo_texto = str(registro.tipo or "").strip()
        tipo_normalizado = tipo_texto.zfill(2) if tipo_texto.isdigit() else tipo_texto
        if tipo_normalizado not in {"03", "05", "06"}:
            raise ValueError(
                f"Tipo de documento inválido (registro {idx + 1})."
            )

        fecha_parseada, fecha_texto = _parse_fecha(
            registro.fecha_emision, periodo, idx
        )

        montos: Dict[str, Decimal] = {}
        for campo in MONTO_FIELDS:
            valor = getattr(registro, campo)
            montos[campo] = _normalize_decimal(valor, campo, idx, False)
        _validar_montos(montos, idx)

        tipo_operacion, tipo_ingreso = _normalize_renta_fields(registro, periodo)

        numero_control = _limpiar_guiones(str(registro.numero_control or "").strip())
        codigo_generacion = _limpiar_guiones(
            str(registro.codigo_generacion or "").strip()
        )
        sello_recepcion = str(registro.sello_recepcion or "").strip()

        limite_noviembre = date(2022, 11, 1)
        if fecha_parseada < limite_noviembre:
            valor_d = codigo_generacion
            valor_f = numero_control
        else:
            valor_d = numero_control
            valor_f = codigo_generacion

        if not valor_d:
            raise ValueError(
                f"Número de control/código de generación faltante para la columna D"
                f" (registro {idx + 1})."
            )
        if not valor_f:
            raise ValueError(
                f"Número de control/código de generación faltante para la columna F"
                f" (registro {idx + 1})."
            )

        identificacion = str(registro.identificacion or "").strip()
        identificacion = _normalizar_identificacion(identificacion) if identificacion else ""
        dui = str(registro.dui or "").strip()
        dui = _normalizar_identificacion(dui) if dui else ""

        nombre = str(registro.nombre_cliente or "").strip()
        nombre = _WHITESPACE_RE.sub(" ", nombre)
        if not nombre:
            raise ValueError(
                f"El nombre o razón social no puede estar vacío (registro {idx + 1})."
            )

        if not identificacion and not dui:
            raise ValueError(
                f"Debe proporcionar NIT/NRC o DUI del cliente (registro {idx + 1})."
            )

        if identificacion and dui:
            raise ValueError(
                "No se puede registrar NIT/NRC y DUI a la vez"
                f" (registro {idx + 1})."
            )

        fila = [
            fecha_texto,
            clase,
            tipo_normalizado,
            valor_d,
            sello_recepcion,
            valor_f,
            "",
            identificacion,
            nombre,
            _decimal_to_text(montos["ventas_exentas"]),
            _decimal_to_text(montos["ventas_no_sujetas"]),
            _decimal_to_text(montos["ventas_gravadas_locales"]),
            _decimal_to_text(montos["debito_fiscal"]),
            _decimal_to_text(montos["ventas_terceros_no_domiciliados"]),
            _decimal_to_text(montos["debito_terceros"]),
            _decimal_to_text(montos["total_ventas"]),
            dui,
            tipo_operacion,
            tipo_ingreso,
            "1",
        ]

        filas.append(((fecha_parseada, tipo_normalizado, valor_d, valor_f), fila))

    filas.sort(key=lambda item: item[0])
    filas_ordenadas = [fila for _, fila in filas]

    csv_name = f"Ventas_Contribuyentes_{periodo}.csv"
    xlsx_name = f"Ventas_Contribuyentes_{periodo}.xlsx"
    csv_path = output_path / csv_name
    xlsx_path = output_path / xlsx_name

    with open(csv_path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(
            fh,
            delimiter=";",
            lineterminator="\n",
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writerows(filas_ordenadas)

    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo I"
    ws.append(HEADERS)
    for fila in filas_ordenadas:
        ws.append(fila)
    _autoajustar_columnas(ws)
    wb.save(xlsx_path)

    return {"csv": csv_path, "xlsx": xlsx_path}


def on_click_generar_contribuyentes(
    output_dir: str | Path,
    periodo_yyyymm: str,
    registros: Iterable[VentaContribuyente],
) -> Dict[str, object]:
    registros_list = list(registros)

    try:
        paths = generar_anexo_contribuyentes_files(
            registros_list, output_dir, periodo_yyyymm
        )
    except ValueError as exc:
        return {
            "success": False,
            "message": str(exc),
            "paths": {},
            "count": 0,
        }
    except Exception as exc:
        return {
            "success": False,
            "message": f"Error inesperado: {exc}",
            "paths": {},
            "count": 0,
        }

    csv_path = paths["csv"]
    xlsx_path = paths["xlsx"]
    cantidad = len(registros_list)
    mensaje = (
        f"{cantidad} DTE exportados al Anexo I. "
        f"CSV: {csv_path}, XLSX: {xlsx_path}."
    )
    return {
        "success": True,
        "message": mensaje,
        "paths": paths,
        "count": cantidad,
    }
